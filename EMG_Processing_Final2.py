from tkinter import *
from tkinter import filedialog
# from tkMessageBox import *
import tkinter

window = Tk()
menu = Menu(window)
d= {}
# for i in enumerate(range(1,17)):
#     d['v{}'.format(i)] = tkinter.IntVar()
for i in range(1,17):
    d['v{}'.format(i)] = tkinter.IntVar()
# for i in range(1,17):
#     print(d['v{}'.format(i)])
# exit()
# v[1] = tkinter.IntVar()
# v[2] = tkinter.IntVar()
# v[3] = tkinter.IntVar()
# v[4] = tkinter.IntVar()
# v[5] = tkinter.IntVar()
# v[6] = tkinter.IntVar()
# v[7] = tkinter.IntVar()
# v[8] = tkinter.IntVar()
# v[9] = tkinter.IntVar()
# v[10] = tkinter.IntVar()
# v[11] = tkinter.IntVar()
# v[12] = tkinter.IntVar()
# v[13] = tkinter.IntVar()
# v[14] = tkinter.IntVar()
# v[15] = tkinter.IntVar()
# v[16] = tkinter.IntVar()



emg_options = [('EMG 1',1),('EMG 2',2),('EMG 3',3),('EMG 4',4),
               ('EMG 5',5),('EMG 6',6),('EMG 7',7),('EMG 8',8),
               ('EMG 9',9),('EMG 10',10),('EMG 11',11),('EMG 12',12),
               ('EMG 13',13),('EMG 14',14),('EMG 15',15),('EMG 16',16)]

# file = filedialog.askopenfilename()
filename = ''
window.title('EMGProc')
# window.geometry('800x600')
csv_checker = 0
color_iter = 0
filenamesplit =[]
emg_choice = []
overwrite_choice = 0
relative_height = 0.8
pulse_width_min = 500
pulse_width_max = 2000
# import EMG_Proc_Func.py


def change_color():
    global color_iter
    global bg_color
    current_color = file_txt.cget("background")
    next_color = bg_color if current_color == "red" else "red"
    file_txt.config(background=next_color)
    if color_iter < 4:
        window.after(500, change_color)
        color_iter += 1
    else:
        color_iter = 0
        return

def dataproc():
    global csv_checker
    global filename
    global filenamesplit
    global emg_choice
    global overwrite_choice
    global output_filename_choice
    global relative_height
    global pulse_width_min
    global pulse_width_max
    if csv_checker == 0:
        from scipy import fftpack, signal
        from matplotlib import pyplot as plt
        import math
        import numpy as np
        import pandas as pd
        import csv
        import scipy.signal as sps
        from scipy.signal import hilbert

        from math import floor, log

        temp = []
        samplerate = 1500
        nyq = samplerate * 0.5
        row_buff_val = 26
        row_buff_val2 = 15

        def _linear_regression(x, y):
            """Fast linear regression using Numba.
            Parameters
            ----------
            x, y : ndarray, shape (n_times,)
                Variables
            Returns
            -------
            slope : float
                Slope of 1D least-square regression.
            intercept : float
                Intercept
            """
            n_times = x.size
            sx2 = 0
            sx = 0
            sy = 0
            sxy = 0
            for j in range(n_times):
                sx2 += x[j] ** 2
                sx += x[j]
                sxy += x[j] * y[j]
                sy += y[j]
            den = n_times * sx2 - (sx ** 2)
            num = n_times * sxy - sx * sy
            slope = num / den
            intercept = np.mean(y) - slope * np.mean(x)
            return slope, intercept

        def _higuchi_fd(x, kmax):
            n_times = x.size
            lk = np.empty(kmax)
            x_reg = np.empty(kmax)
            y_reg = np.empty(kmax)
            for k in range(1, kmax + 1):
                lm = np.empty((k,))
                for m in range(k):
                    ll = 0
                    n_max = floor((n_times - m - 1) / k)
                    n_max = int(n_max)
                    for j in range(1, n_max):
                        ll += abs(x[m + j * k] - x[m + (j - 1) * k])
                    ll /= k
                    ll *= (n_times - 1) / (k * n_max)
                    lm[m] = ll
                # Mean of lm
                m_lm = 0
                for m in range(k):
                    m_lm += lm[m]
                m_lm /= k
                lk[k - 1] = m_lm
                x_reg[k - 1] = log(1. / k)
                y_reg[k - 1] = log(m_lm)
            higuchi, _ = _linear_regression(x_reg, y_reg)
            return higuchi

        from openpyxl import Workbook
        from openpyxl.chart import (ScatterChart, Reference, Series)
        import openpyxl

        # dest_filename = 'Subj08.xlsx'
        if len(output_filename_choice.get())==0:
            dest_filename = filename +'.xlsx'
        else:
            dest_filename = output_filename_choice.get()+'.xlsx'

        if overwrite_choice == 0:
            book = Workbook()
        else:
            book = openpyxl.load_workbook(dest_filename)
        sheet = book.active
        sheet.title = 'Data'

        sheet2 = book.create_sheet()
        sheet2.title = 'Graphs'

        with open(filename, 'r') as csvfile:
            csvreader = csv.reader(csvfile, delimiter=',')
            for row in csvreader:
                if csvreader.line_num == 3:
                    temp.append(row)
                if csvreader.line_num >= 6:
                    if row:
                        temp.append(row)
                    else:
                        break

        df = pd.DataFrame(temp)  # turns the array into a dataframe
        df.columns = df.iloc[0]  # sets the column names as the first row
        df = df.drop(0)  # drops the first row since it is now a duplicate of the column names
        df.reindex(df.index.drop(1))
        df.reset_index(drop=True, inplace=True)
        df.columns = ['frames', 'subframes', 'blank', 'emg1', 'emg2', 'emg3', 'emg4', 'emg5', 'emg6', 'emg7', 'emg8',
                      'emg9', 'emg10', 'emg11', 'emg12', 'emg13', 'emg14', 'emg15', 'emg16', 'unused7', 'unused8',
                      'blank2']
        df2 = df.drop(['frames', 'subframes', 'blank', 'unused7', 'unused8', 'blank2'], axis=1)
        df2 = df2.astype(np.float)
        print(len(df))
        hor = np.arange(0, (len(df) - 0.5) / samplerate, 1 / samplerate)  # getting the time domain in seconds

        emg = []
        for i in range(0,len(emg_choice)):
            j = emg_choice[i]
            print(j)
            emg.append(df2['emg{}'.format(j)])

        # print(emg[0])
        # print('blank')
        # print(len(emg))
        # print(emg[1][3])
        # exit()

        cutoff_freq = 20  # ~20 Hz for movement according to emg book "Electromyography: Physiology, Engineering, and Noninvasive Apps"
        cut = cutoff_freq / nyq
        b, a = signal.butter(5, cut, btype='highpass', analog=False)
        emg_high = []
        for i in range(0, len(emg)):
            temp = signal.filtfilt(b, a, emg[i])
            emg_high.append(temp)
        # print(emg_high[0])
        # print(len(emg_high))
        # exit()

        cutoff_freq = 400  # ~500 Hz according to the emg book
        cut = cutoff_freq / nyq
        b, a = signal.butter(5, cut, btype='lowpass', analog=False)
        emg_filt = []
        for i in range(0, len(emg_high)):
            temp = signal.filtfilt(b, a, emg_high[i])
            emg_filt.append(temp)

        # plt.figure(1)
        # plt.subplot(2, 4, 1)
        # plt.plot(hor, emg1)
        # plt.subplot(2, 4, 2)
        # plt.plot(hor, emg2)
        # plt.subplot(2, 4, 3)
        # plt.plot(hor, emg3)
        # plt.subplot(2, 4, 4)
        # plt.plot(hor, emg4)
        # plt.subplot(2, 4, 5)
        # plt.plot(hor, emg5)
        # plt.subplot(2, 4, 6)
        # plt.plot(hor, emg6)
        # plt.subplot(2, 4, 7)
        # plt.plot(hor, emg7)

        # plt.show()
        emg_rec = []
        for i in range(0, len(emg_filt)):
            temp = abs(emg_filt[i])
            emg_rec.append(temp)


        ynew = []
        for i in range(0, len(emg_rec)):
            temp = signal.savgol_filter(emg_rec[i], 1501, 2)
            ynew.append(temp)


        cutoff_freq = 10  # ~20 Hz for movement according to emg book "Electromyography: Physiology, Engineering, and Noninvasive Apps"
        cut = cutoff_freq / nyq
        b, a = signal.butter(5, cut, btype='lowpass', analog=False)

        ynew2 = []
        for i in range(0, len(ynew)):
            temp = signal.filtfilt(b, a, ynew[i])
            ynew2.append(temp)


        ynew3 = []
        for i in range(0, len(ynew2)):
            temp = signal.savgol_filter(ynew2[i], 1501, 2)
            ynew3.append(temp)

        ynew3 = []
        for i in range(0, len(ynew2)):
            temp = signal.savgol_filter(ynew2[i], 1501, 2)
            ynew3.append(temp)
        peaks = []
        props = []
        for i in range(0, len(ynew3)):
            temp1,temp2 = signal.find_peaks(ynew3[i], width=(pulse_width_min, pulse_width_max), rel_height=relative_height)
            peaks.append(temp1)
            props.append(temp2)

        # print("len of peaks1: " + str(len(peaks1)))

        pulses_beginT = [[] for _ in range(len(peaks))]
        pulses_endT = [[] for _ in range(len(peaks))]
        pulses_begin = [[] for _ in range(len(peaks))]
        pulses_end = [[] for _ in range(len(peaks))]
        pulses_begin_ind = [[] for _ in range(len(peaks))]
        pulses_end_ind = [[] for _ in range(len(peaks))]
        # print(props[0]['widths'])
        # print(peaks[0])
        print(len(peaks))
        print('blank')
        # exit()
        for i in range(len(peaks)):
            for j,k in zip(peaks[i], props[i]['widths']):
                # for l in len(j):
                pulse_sample_start = j - (math.floor(k / 2))
                pulse_sample_end = j + (math.floor(k / 2))
                pulses_begin_ind[i].append(pulse_sample_start)
                pulses_end_ind[i].append(pulse_sample_end)

                pulses_beginT[i].append(pulse_sample_start / 1500)
                pulses_endT[i].append(pulse_sample_end / 1500)
                pulses_begin[i].append(ynew3[i][pulse_sample_start])
                pulses_end[i].append(ynew3[i][pulse_sample_end])

                # print(j)
                # print(k)
            # print(pulses_beginT[0])
            # print(pulses_endT[0])
            # print(pulses_begin[0])
            # print(pulses_end[0])
            # print(pulses_begin_ind[0])
            # print(pulses_end_ind[0])

            # exit()

        sectionpoints = []
        sectionpoints_array = []
        sectionT = []
        # row = 2
        lowcut = 10
        highcut = 400
        for t in range(len(ynew3)):
            M = 0
            N = 1
            cellcheck = 0
            # print('double blank')
            # print(len(peaks[t]))
            # print(t)
            # print(emg_rec[0])
            # for a in range((len(peaks[t]))):

                # for i in range(M, N):
                #     for q in emg_rec[t]:
            for r,s in zip(pulses_begin_ind[t],pulses_end_ind[t]):
                # print('in for loop')
                # print(a)
                # print(pulses_begin_ind[t])
                # print(pulses_end_ind[t])
                # print(r)
                # print(s)
                # exit()
                q = emg_rec[t]
                sectionpoints = []
                sectionT = []
                sectionpoints.extend(q[r:s])
                sectionT.extend(hor[r:s])
                # print(a)
                # exit()
                # x1 = np.linspace(0, len(sectionpoints1), len(sectionpoints1))
                # x1 = x1 / 1500

                sectionpoints_array = np.asarray(sectionpoints)
                freq, power_spec = signal.periodogram(sectionpoints_array, samplerate)

                lowflag = 0
                breakflag = 0
                # for c in range(len(freq[t])):
                for c in range(len(freq)):
                    if abs(freq[c]) > 10 and lowflag == 0:
                        fullpulse_lowfreqbound = c
                        lowflag = 1
                    if abs(freq[c]) > 400:
                        fullpulse_uppfreqbound = c
                        break

                spec_mom0 = 0
                spec_mom2 = 0
                spec_mom5 = 0

                for k in range(fullpulse_lowfreqbound, fullpulse_uppfreqbound):
                    spec_mom0 = spec_mom0 + (math.pow(freq[k], -1) * power_spec[k])
                    spec_mom2 = spec_mom2 + (math.pow(freq[k], 2) * power_spec[k])
                    spec_mom5 = spec_mom5 + (math.pow(freq[k], 5) * power_spec[k])

                f2 = spec_mom0 / spec_mom2
                f5 = spec_mom0 / spec_mom5

                powsum = 0
                powarray = []

                for l in range(len(freq)):
                    powsum = powsum + power_spec[l]
                    powarray.append(powsum)

                mednum = powsum / 2

                meansumcombo = 0
                meansumpow = 0

                for p in range(len(freq)):
                    meansumcombo = meansumcombo + (freq[p] * power_spec[p])
                    meansumpow = meansumpow + (power_spec[p])

                mean = meansumcombo / meansumpow

                row = 2 + t*row_buff_val
                col = 1
                cellcheck = 0
                for u in powarray:
                    if u > mednum:
                        median = freq[powarray.index(u)]
                        while cellcheck == 0:
                            if sheet.cell(row=row, column=col).value != None:
                                col = col + 1
                            else:
                                cellcheck = 1
                        sheet.cell(row=row, column=col).value = median
                        break

                cellcheck = 0
                row = row + 2
                col = 1
                while cellcheck == 0:
                    if sheet.cell(row=row, column=col).value != None:
                        col = col + 1
                    else:
                        cellcheck = 1
                sheet.cell(row=row, column=col).value = mean

                cellcheck = 0
                row = row + 2
                col = 1
                while cellcheck == 0:
                    if sheet.cell(row=row, column=col).value != None:
                        col = col + 1
                    else:
                        cellcheck = 1
                sheet.cell(row=row, column=col).value = f2

                cellcheck = 0
                row = row + 2
                col = 1
                while cellcheck == 0:
                    if sheet.cell(row=row, column=col).value != None:
                        col = col + 1
                    else:
                        cellcheck = 1
                sheet.cell(row=row, column=col).value = f5



            row = row + 2
            col = 1
            cellcheck = 0
            for v in range(len(emg_filt[t]) // 750):
                q2 = emg_filt[t]
                emg_sec = q2[(0 + (750 * v)):(7500 + (750 * v))]
                frac1 = _higuchi_fd(emg_sec, 6)
                while cellcheck == 0:
                    if sheet.cell(row=row, column=col).value != None:
                        col = col + 1
                    else:
                        cellcheck = 1
                sheet.cell(row=row, column=col).value = frac1
                col = col + 1

            # row = row + 5
            # col = 1
            cellcheck = 0


        # ^row 92, 94, 96, 98, 100, 102
        col1 = []
        col2 = []
        for e in range(len(ynew3)):
            row = 2 + e*row_buff_val
            # print(row)
            col1.append(1)
            # print(col1[e])
            cellcheck = 0
            while cellcheck == 0:
                if sheet.cell(row=row, column=col1[e]).value != None:
                    sheet.cell(row=(row-1), column=col1[e]).value = col1[e]
                    col1[e] = col1[e] + 1
                else:
                    cellcheck = 1

            row = 10 + e*row_buff_val
            col2.append(1)
            cellcheck = 0
            while cellcheck == 0:
                if sheet.cell(row=row, column=col2[e]).value != None:
                    sheet.cell(row=(row-1), column=col2[e]).value = col2[e]
                    col2[e] = col2[e] + 1
                else:
                    cellcheck = 1
        # print(col1)
        # print(col2)
        # exit()
        # chart = ScatterChart()
        # chart.title = "EMG 1 Median Frequency"
        # # chart.style = 13
        # chart.x_axis.title = 'Pulse'
        # chart.y_axis.title = 'Frequency'


        # for i in range(0,len(emg_choice)):  #just for reference for charts
        #     j = emg_choice[i]

        for b in range(len(emg_choice)):
            for c in range(1,6):
                # print('c = ' +str(c))
                chart = ScatterChart()
                # if row == (2 + b * row_buff_val):
                if c == 1:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Median Frequency"
                # elif row == (4 + b * row_buff_val):
                elif c == 2:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Mean Frequency"
                # elif row == (6 + b * row_buff_val):
                elif c == 3:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Spectral Index Order 2"
                # elif row == (8 + b * row_buff_val):
                elif c == 4:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Spectral Index Order 5"
                # elif row == (10 + b * row_buff_val):
                elif c == 5:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Fractal Dimension"

                # chart.style = 13
                if c == 5:
                    chart.x_axis.title = 'Window'
                else:
                    chart.x_axis.title = 'Pulse'

                if c == 1 or c == 2:
                    chart.y_axis.title = 'Frequency'
                elif c == 3 or c == 4:
                    chart.y_axis.title = 'Spectral Index Power'
                else:
                    chart.y_axis.title = 'Fractal Dimension'

                if c == 5:
                    xvalues = Reference(sheet, min_col=1, max_col = (col2[b]-1), min_row = 9+b*row_buff_val)
                else:
                    xvalues = Reference(sheet, min_col=1, max_col = (col1[b]-1), min_row = 1+b*row_buff_val)

                if c == 5:
                    values = Reference(sheet, min_col =1, max_col = (col2[b]-1), min_row = (10 + b*row_buff_val))
                else:
                    values = Reference(sheet, min_col=1, max_col=(col1[b]-1), min_row=((2*c)+(b*row_buff_val)))

                series = Series(values, xvalues)
                chart.series.append(series)
                chartloc = ''
                if c == 1:
                    chartloc = 'A'
                elif c == 2:
                    chartloc = 'J'
                elif c == 3:
                    chartloc = 'S'
                elif c == 4:
                    chartloc = 'AB'
                elif c == 5:
                    chartloc = 'AK'
                # print(chartloc)
                # print(chartloc + str(11+b*row_buff_val))
                sheet.add_chart(chart, chartloc + str(11+b*row_buff_val))
                # sheet2.add_chart(chart, chartloc + str(11 + b * row_buff_val))
                # exit()

        for b in range(len(emg_choice)):
            for c in range(1,6):
                # print('c = ' +str(c))
                chart = ScatterChart()
                # if row == (2 + b * row_buff_val):
                if c == 1:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Median Frequency"
                # elif row == (4 + b * row_buff_val):
                elif c == 2:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Mean Frequency"
                # elif row == (6 + b * row_buff_val):
                elif c == 3:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Spectral Index Order 2"
                # elif row == (8 + b * row_buff_val):
                elif c == 4:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Spectral Index Order 5"
                # elif row == (10 + b * row_buff_val):
                elif c == 5:
                    chart.title = "EMG "+ str(emg_choice[b]) + " Fractal Dimension"

                # chart.style = 13
                if c == 5:
                    chart.x_axis.title = 'Window'
                else:
                    chart.x_axis.title = 'Pulse'

                if c == 1 or c == 2:
                    chart.y_axis.title = 'Frequency'
                elif c == 3 or c == 4:
                    chart.y_axis.title = 'Spectral Index Power'
                else:
                    chart.y_axis.title = 'Fractal Dimension'

                if c == 5:
                    xvalues = Reference(sheet, min_col=1, max_col = (col2[b]-1), min_row = 9+b*row_buff_val)
                else:
                    xvalues = Reference(sheet, min_col=1, max_col = (col1[b]-1), min_row = 1+b*row_buff_val)

                if c == 5:
                    values = Reference(sheet, min_col =1, max_col = (col2[b]-1), min_row = (10 + b*row_buff_val))
                else:
                    values = Reference(sheet, min_col=1, max_col=(col1[b]-1), min_row=((2*c)+(b*row_buff_val)))

                series = Series(values, xvalues)
                chart.series.append(series)
                chartloc = ''
                if c == 1:
                    chartloc = 'A'
                elif c == 2:
                    chartloc = 'J'
                elif c == 3:
                    chartloc = 'S'
                elif c == 4:
                    chartloc = 'AB'
                elif c == 5:
                    chartloc = 'AK'
                # print(chartloc)
                # print(chartloc + str(11+b*row_buff_val))
                sheet2.add_chart(chart, chartloc + str(11+b*row_buff_val2))
                # sheet2.add_chart(chart, chartloc + str(11 + b * row_buff_val))

        # count = 0
        # while sheet.cell(row=2, column=count+1).value != None:
        #     count = count+1
        # xvalues = list(range(count))
        # yval = Reference(sheet, min_col = 1, min_row = 2, max_col = count, max_row =2)
        # series = Series(values=yval, xvalues=xval)
        #
        # chart.series.append(series)
        #
        # sheet.add_chart(chart, 'A10')

        # chart.title = "EMG 1 Mean Frequency"
        #
        # chart.title = "EMG 1 Spectral Moment Order 2"
        #
        # chart.title = "EMG 1 Spectral Moment Order 5"
        #
        # chart.title = "EMG 1 Fractal Dimension"
        # exit()
        print(dest_filename)
        book.save(filename=dest_filename)

        # fig = plt.figure(2)
        # ax0 = fig.add_subplot(2, 4, 1)
        # ax0.plot(hor, emg_rec1)
        # ax0.plot(hor, ynew1, color='red')
        # ax0.plot(pulses_beginT1, pulses_begin1, marker="x", color='yellow', linestyle="None")
        # ax0.plot(pulses_endT1, pulses_end1, marker="x", color='yellow', linestyle="None")
        #
        # ax0 = fig.add_subplot(2, 4, 2)
        # ax0.plot(hor, emg_rec2)
        # ax0.plot(hor, ynew2, color='red')
        # ax0.plot(pulses_beginT2, pulses_begin2, marker="x", color='yellow', linestyle="None")
        # ax0.plot(pulses_endT2, pulses_end2, marker="x", color='yellow', linestyle="None")
        #
        # ax0 = fig.add_subplot(2, 4, 3)
        # ax0.plot(hor, emg_rec3)
        # ax0.plot(hor, ynew3, color='red')
        # ax0.plot(pulses_beginT3, pulses_begin3, marker="x", color='yellow', linestyle="None")
        # ax0.plot(pulses_endT3, pulses_end3, marker="x", color='yellow', linestyle="None")
        #
        # ax0 = fig.add_subplot(2, 4, 4)
        # ax0.plot(hor, emg_rec4)
        # ax0.plot(hor, ynew4, color='red')
        # ax0.plot(pulses_beginT4, pulses_begin4, marker="x", color='yellow', linestyle="None")
        # ax0.plot(pulses_endT4, pulses_end4, marker="x", color='yellow', linestyle="None")
        #
        # ax0 = fig.add_subplot(2, 4, 5)
        # ax0.plot(hor, emg_rec5)
        # ax0.plot(hor, ynew5, color='red')
        # ax0.plot(pulses_beginT5, pulses_begin5, marker="x", color='yellow', linestyle="None")
        # ax0.plot(pulses_endT5, pulses_end5, marker="x", color='yellow', linestyle="None")
        #
        # ax0 = fig.add_subplot(2, 4, 6)
        # ax0.plot(hor, emg_rec6)
        # ax0.plot(hor, ynew6, color='red')
        # ax0.plot(pulses_beginT6, pulses_begin6, marker="x", color='yellow', linestyle="None")
        # ax0.plot(pulses_endT6, pulses_end6, marker="x", color='yellow', linestyle="None")
        #
        # ax0 = fig.add_subplot(2, 4, 7)
        # ax0.plot(hor, emg_rec7)
        # ax0.plot(hor, ynew7, color='red')
        # ax0.plot(pulses_beginT7, pulses_begin7, marker="x", color='yellow', linestyle="None")
        # ax0.plot(pulses_endT7, pulses_end7, marker="x", color='yellow', linestyle="None")

        # plt.show()

        gen_txt.configure(text='Generation Complete!', background = 'green')
        # gen_txt.grid(row = 20, column = 2, columnspan = 3)
    else:
        change_color()


def filebrowse():
    global csv_checker
    global filename
    global filenamesplit
    gen_txt.configure(text='', background=bg_color)
    output_filename_choice.config(state='normal')
    PulseRelHeight.config(state='normal')
    PulseWidthMin.config(state='normal')
    PulseWidthMax.config(state='normal')
    filename = filedialog.askopenfilename()

    if filename.endswith('.csv'):
        filenamesplit = filename.split('/')
        file_txt.configure(text='Chosen File: ' + filenamesplit[-1])
        csv_checker = 0
        file_txt.config(background=bg_color)
    else:
        file_txt.configure(text='Please choose a csv file!')
        csv_checker = 1

def get_emg_vals():
    global emg_choice
    # print(emg_choice)
    emg_choice=[]
    for i in range(1,17):
        # print(i)
        # print(d['v{}'.format(i)].get())
        if d['v{}'.format(i)].get() == 1:
            emg_choice.append(i)
    print(emg_choice)

def NewFile():
    global emg_choice
    global CB1
    # global file_txt
    file_txt.configure(text='Chosen File: ')
    for i in range(1,17):
        d['v{}'.format(i)].set(0)
    gen_txt.configure(text='', background=bg_color)

def NewFileKeyboard(self):
    global emg_choice
    global CB1
    # global file_txt
    file_txt.configure(text='Chosen File: ')
    for i in range(1, 17):
        d['v{}'.format(i)].set(0)
    gen_txt.configure(text='', background=bg_color)

def ExitKeyboard(self):
    window.quit()


def Tutorial():
    top = Toplevel()
    top.title("Tutorial")
    # top.geometry('500x400')
    msg1 = Label(top, text = 'Hello!'
                            '\n Welcome to EMGProc!'
                            '\nThis software was designed to work with a 16 count EMG system, as well as'
                            '\n the csv file format created by Vicon Nexus (specifically v2.5 but possibly could'
                            '\n work for other versions)'
                            '\n'
                            '\n'
                            '\n How to use EMGProc:')
    msg1.grid(row=1,column = 1,pady = (20,0),padx = 30)

    msg2 = Label(top, text ='Step 1: Choose the csv file needed to process'
                            '\nStep 2: Choose the output filename, as well as if a new output file should be created'
                            '\n(overwriting any file with that name) or add on to an existing file created by EMGProc'
                            '\nStep 3: Choose what EMGs were active and save them'
                            '\nStep 4: Click Generate Output'
                            '\nThe output file will be placed in the same location as the chosen file, unless you choose,'
                            '\nyour own filename, in which case it will be placed in the location of this program.'
                            '\n', justify = LEFT)
    msg2.grid(row=2,column = 1,padx = 30)
    msg3 = Label(top, text ='Keyboard Shortucts:')
    msg3.grid(row=3,column = 1,padx = 30)
    msg4 = Label(top, text='Ctrl+Q: New File'
                            '\nCtrl+X: Exit', justify = LEFT)
    msg4.grid(row=4,column = 1,pady = (0,20))

    button1 = Button(top, text = 'Dismiss', command = top.destroy)
    button1.grid(row=5,column = 1,pady = (0,20))

def About():
    top = Toplevel()
    top.title("About")
    # top.geometry('500x400')
    msg1 = Label(top, text = 'EMGProc was created by Adam Lewis as part of a Master\'s Thesis Project'
                             '\nCreated with PyCharm IDE and Python 3.6'
                             '\nContact email: AdamLew73@gmail.com'
                             '\n'
                             '\nPlease do not copy or use without permission')
    msg1.grid(row=1,column = 1,pady = (20,0),padx = 30)


    button1 = Button(top, text='Dismiss', command=top.destroy)
    button1.grid(row=2, column=1, pady=(0, 20))

def focus_out_entry(event):
    output_filename.config(text='\nOutput Filename: ' + output_filename_choice.get()+'.xlsx')
def output_filename_save():
    output_filename.config(text='\nOutput Filename: ' + output_filename_choice.get() + '.xlsx')

def save_choice1():
    global overwrite_choice
    overwrite_choice = 0
    print(overwrite_choice)
def save_choice2():
    global overwrite_choice
    overwrite_choice = 1
    print(overwrite_choice)
def Pulse_Prop():
    global relative_height
    global pulse_width_min
    global pulse_width_max
    relative_height = float(PulseRelHeight.get())
    pulse_width_min = int(PulseWidthMin.get())
    pulse_width_max = int(PulseWidthMax.get())

##### GUI CREATION #####
window.config(menu = menu)

introtext = Label(window, text = '---------------------------------------------------------------------------------'
                               '\nFile Management')
introtext.grid(row=1,column = 2, columnspan = 3)

btn1 = Button(window, text = 'Browse for File (csv only)', command = filebrowse)
# btn1.pack(fill=X,padx = 100, pady = 50)
btn1.grid(row=2,column = 3,pady = 20)

file_txt = Label(window, text = 'Chosen File: ' + filename)
# file_txt.pack(fill=X)
file_txt.grid(row=3,column = 2, columnspan = 3, ipadx = 100)
bg_color = file_txt.cget("background")
out_text = ''
output_filename_txt = Label(window, text = 'Desired Output Filename:')
output_filename_txt.grid(row=4,column = 2, columnspan =2,padx = (25,100), sticky = W)
output_filename_choice = Entry(window, textvariable = out_text, state = 'disabled', width = 50)
output_filename_choice.bind("<Return>",focus_out_entry)
output_filename_choice.bind("<FocusOut>",focus_out_entry)
output_filename_choice.grid(row=4,column = 2, columnspan = 2, sticky = W, padx = (175,0))
output_file_save = Button(window, text = 'Enter', command = output_filename_save)
output_file_save.grid(row=4,column = 4, padx = (10,0), sticky = W)
filesavechoicevar = IntVar()
file_save_choice1 = Radiobutton(window, text = 'Overwrite File', variable = filesavechoicevar, value = 0, command = save_choice1)
file_save_choice1.grid(row=3,column = 4, padx = (50,25), pady = (10,30), rowspan = 3, sticky = NE)
file_save_choice2 = Radiobutton(window, text = 'Add to File', variable = filesavechoicevar, value = 1, command = save_choice2)
file_save_choice2.grid(row=3,column = 4, padx = (0,40), pady = (35,0), rowspan = 3, sticky = NE)

output_filename = Label(window, text = '\nOutput Filename: ')
output_filename.grid(row=5,column = 3)
septext = Label(window, text = '---------------------------------------------------------------------------------'
                               '\nActive EMGs'
                               '\n')
septext.grid(row=6,column = 2, columnspan = 3)

CB1 =Checkbutton(window, text = 'EMG 1', variable = d['v1'])
CB1.grid(row=7, column = 2, ipadx = 50)
CB2 =Checkbutton(window, text = 'EMG 2', variable = d['v2'])
CB2.grid(row=8, column = 2)
CB3 =Checkbutton(window, text = 'EMG 3', variable = d['v3'])
CB3.grid(row=9, column = 2)
CB4 =Checkbutton(window, text = 'EMG 4', variable = d['v4'])
CB4.grid(row=10, column = 2)
CB5 =Checkbutton(window, text = 'EMG 5', variable = d['v5'])
CB5.grid(row=11, column = 2)
CB6 =Checkbutton(window, text = 'EMG 6', variable = d['v6'])
CB6.grid(row=12, column = 2)
CB7 =Checkbutton(window, text = 'EMG 7', variable = d['v7'])
CB7.grid(row=7, column = 3)
CB8 =Checkbutton(window, text = 'EMG 8', variable = d['v8'])
CB8.grid(row=8, column = 3)
CB9 =Checkbutton(window, text = 'EMG 9', variable = d['v9'])
CB9.grid(row=9, column = 3)
CB10 =Checkbutton(window, text = 'EMG 10', variable = d['v10'])
CB10.grid(row=10, column = 3)
CB11 =Checkbutton(window, text = 'EMG 11', variable = d['v11'])
CB11.grid(row=11, column = 3)
CB12 =Checkbutton(window, text = 'EMG 12', variable = d['v12'])
CB12.grid(row=7, column = 4, ipadx = 50)
CB13 =Checkbutton(window, text = 'EMG 13', variable = d['v13'])
CB13.grid(row=8, column = 4)
CB14 =Checkbutton(window, text = 'EMG 14', variable = d['v14'])
CB14.grid(row=9, column = 4)
CB15 =Checkbutton(window, text = 'EMG 15', variable = d['v15'])
CB15.grid(row=10, column = 4)
CB16 =Checkbutton(window, text = 'EMG 16', variable = d['v16'])
CB16.grid(row=11, column = 4)

btn2 = Button(window, text = 'Save Active EMGs', command = get_emg_vals)
btn2.grid(row=13,column = 3,pady = 10)
# file_txt = Label(window, text = 'Chosen File: ' + filename)
# file_txt.pack(fill=X)

septext2 = Label(window, text = '---------------------------------------------------------------------------------'
                               '\nPulse Definitions'
                               '\n')
septext2.grid(row=14,column = 2, columnspan = 3)

v1 = StringVar(window, value='0.8')
v2 = StringVar(window, value='500')
v3 = StringVar(window, value='2000')

PulseRelHeight_Text = Label(window, text = 'Pulse Relative Height')
PulseRelHeight_Text.grid(row=15,column = 2, columnspan = 2, padx = (75,100), sticky = W)
PulseRelHeight = Entry(window, state = 'disabled', textvariable = v1, width = 10)
PulseRelHeight.grid(row=16,column = 2, columnspan = 2, padx = (100,100), sticky = W)

PulseWidthMin_Text = Label(window, text = 'Pulse Min Width(in Samples)')
PulseWidthMin_Text.grid(row=15,column = 2, columnspan = 3, padx = (50,50))
PulseWidthMin = Entry(window, state = 'disabled', textvariable = v2, width = 10)
PulseWidthMin.grid(row=16,column = 2, columnspan = 3, padx = (45,50))

PulseWidthMax_Text = Label(window, text = 'Pulse Max Width(in Samples)')
PulseWidthMax_Text.grid(row=15,column = 3, columnspan = 2, padx = (50,50), sticky = E)
PulseWidthMax = Entry(window, state = 'disabled', textvariable = v3, width = 10)
PulseWidthMax.grid(row=16,column = 3, columnspan = 2, padx = (25,100), sticky = E)

Pulse_Button = Button(window, text = 'Save Properties', command = Pulse_Prop)
Pulse_Button.grid(row=17,column = 3,pady = 20)

septext3 = Label(window, text = '---------------------------------------------------------------------------------')
septext3.grid(row=18,column = 2, columnspan = 3)

btn3 = Button(window, text = 'Generate Output', command = dataproc)
btn3.grid(row=19,column = 3)
gen_txt = Label(window, text = '')
gen_txt.grid(row=20,column = 3, ipadx = 150, pady = (10,20))
# gen_txt.grid(row=20,column = 2,columnspan = 3, ipadx = 150, pady = (10,20))

# btn2.grid(fill=X,padx = 100, pady = 220)


filemenu = Menu(menu, tearoff = False)
filemenu2 = Menu(menu, tearoff = False)
filemenu3 = Menu(menu, tearoff = False)

menu.add_cascade(label='File', menu = filemenu)
filemenu.add_command(label='New', command=NewFile, accelerator ="Ctrl+Q")
window.bind_all("<Control-q>",NewFileKeyboard)
filemenu.add_separator()
filemenu.add_command(label='Exit', command=window.quit)
window.bind_all("<Control-x>",ExitKeyboard)
menu.add_cascade(label='Help', menu = filemenu2)
filemenu2.add_command(label='Tutorial', command=Tutorial)

menu.add_cascade(label='Info', menu = filemenu3)
filemenu3.add_command(label='About', command=About)


window.mainloop()