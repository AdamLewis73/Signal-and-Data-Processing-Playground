# import xlwt
# from xlwt import Workbook
#
# wb = Workbook()
#
# sheet1 = wb.add_sheet('Sheet 1')
# sheet1.write(1,0, 'hi')
# sheet1.write(2,0, 'it works')
# sheet1.write(0,1, 'please work')
# sheet1.write(0,2, 'I is scurd')
#
# wb.save('xlwt example.xls')

# from openpyxl import Workbook
# # import openpyxl
#
# # dest_filename = 'C:\Users\sword\PycharmProjects\excelimport\xlwt example.xlsx'
# # dest_filename = 'xlwt example.xlsx'
#
# # book = openpyxl.load_workbook(dest_filename)
# # sheet = book.active
# data = [('Nooo','Please Work','why does life suck')]
# for x in data:
#     sheet.append(x)
# # for row in range(2,10):
# #     for col in range(1,2):
# #         ws1.cell(column= col, row = row, value = row)
# #         ws1.
# book.save(filename=dest_filename)






### CSV Extraction Code with help from: Jim Todd of Stack Overflow
from scipy import fftpack, signal
from matplotlib import pyplot as plt
import math
import numpy as np
import pandas as pd
import csv
import scipy.signal as sps
from scipy.signal import hilbert

from math import floor,log
# from .utils import _linear_regression

temp = []
samplerate = 1500
nyq = samplerate*0.5

from openpyxl import Workbook
import openpyxl


dest_filename = 'xlwt example.xlsx'
# book = Workbook()
book = openpyxl.load_workbook(dest_filename)
sheet = book.active

# filename = 'C:/Users/sword/Anaconda3/envs/exceltest/RF_Subj08/Free/RF_Subj08_Free_Fatigue5_TR01.csv'
filename = 'C:/Users/sword/Anaconda3/envs/exceltest/RF_Subj08/Free/RF_Subj08_Free_Fatigue5_TR24.csv'
import openpyxl

with open(filename, 'r') as csvfile:
    csvreader = csv.reader(csvfile, delimiter=',')
    for row in csvreader:
        if csvreader.line_num == 3:
            temp.append(row)
        if csvreader.line_num >= 6:
            if row:
                temp.append(row)
            else:
                break

df = pd.DataFrame(temp)  # turns the array into a dataframe
df.columns = df.iloc[0]  # sets the column names as the first row
df = df.drop(0)  # drops the first row since it is now a duplicate of the column names
df.reindex(df.index.drop(1))
df.reset_index(drop=True, inplace=True)
# print(df)
df.columns = ['frames', 'subframes', 'blank', 'emg1', 'emg2', 'emg3', 'emg4', 'emg5', 'emg6', 'emg7', 'emg8', 'emg9',
              'emg10', 'emg11', 'emg12', 'emg13', 'emg14', 'emg15', 'emg16', 'unused7', 'unused8', 'blank2']
# df.columns = ['frames', 'subframes', 'blank', 'NU1', 'NU2', 'NatTA', 'NU3', 'NatBic', 'NatTri', 'NatGastLat', 'NatGastMed', 'EmilyBic', 'EmilyTri', 'EmilyGastMed', 'EmilyGastLat', 'EmilyTA', 'AdamBic', 'AdamTri', 'unused6', 'unused7', 'unused8', 'blank2']
# df.columns = ['frames', 'subframes', 'blank', 'RGastLat', 'LGastLat', 'RTibAnt', 'LTibAnt', 'RBicFem', 'LBicFem', 'RRecFem', 'LRecFem', 'RGastMed', 'LGastMed', 'unused1', 'unused2', 'unused3', 'unused4', 'unused5', 'unused6', 'unused7', 'unused8', 'blank2']
# df.columns = ['frames', 'subframes', 'blank', 'RGastLat', 'LGastLat', 'RTibAnt', 'LTibAnt', 'RBicFem', 'LBicFem', 'RRecFem', 'LRecFem', 'RGastMed', 'LGastMed', 'unused1', 'unused2', 'unused3', 'unused7', 'unused8', 'blank2']
df2 = df.drop(['frames', 'subframes', 'blank', 'unused7', 'unused8', 'blank2'], axis=1)
# print(df2)
df2 = df2.astype(np.float)
print(len(df))
hor = np.arange(0, (len(df) - 0.5) / samplerate, 1 / samplerate)  # getting the time domain in seconds
emg1 = df2.emg1

cutoff_freq = 20  # ~20 Hz for movement according to emg book "Electromyography: Physiology, Engineering, and Noninvasive Apps"
cut = cutoff_freq/nyq
b, a = signal.butter(5, cut, btype='highpass', analog=False)
emg1high = signal.filtfilt(b, a, emg1)

cutoff_freq = 400  # ~500 Hz according to the emg book
cut = cutoff_freq/nyq
b, a = signal.butter(5, cut, btype='lowpass', analog=False)
emg1filt = signal.filtfilt(b, a, emg1high)

def _linear_regression(x, y):
    """Fast linear regression using Numba.
    Parameters
    ----------
    x, y : ndarray, shape (n_times,)
        Variables
    Returns
    -------
    slope : float
        Slope of 1D least-square regression.
    intercept : float
        Intercept
    """
    n_times = x.size
    sx2 = 0
    sx = 0
    sy = 0
    sxy = 0
    for j in range(n_times):
        sx2 += x[j] ** 2
        sx += x[j]
        sxy += x[j] * y[j]
        sy += y[j]
    den = n_times * sx2 - (sx ** 2)
    num = n_times * sxy - sx * sy
    slope = num / den
    intercept = np.mean(y) - slope * np.mean(x)
    return slope, intercept

def _higuchi_fd(x, kmax):
    """Utility function for `higuchi_fd`.
    """
    n_times = x.size
    lk = np.empty(kmax)
    x_reg = np.empty(kmax)
    y_reg = np.empty(kmax)
    for k in range(1, kmax + 1):
        lm = np.empty((k,))
        for m in range(k):
            ll = 0
            n_max = floor((n_times - m - 1) / k)
            n_max = int(n_max)
            for j in range(1, n_max):
                ll += abs(x[m + j * k] - x[m + (j - 1) * k])
            ll /= k
            ll *= (n_times - 1) / (k * n_max)
            lm[m] = ll
        # Mean of lm
        m_lm = 0
        for m in range(k):
            m_lm += lm[m]
        m_lm /= k
        lk[k - 1] = m_lm
        x_reg[k - 1] = log(1. / k)
        y_reg[k - 1] = log(m_lm)
    higuchi, _ = _linear_regression(x_reg, y_reg)
    return higuchi

# sheet.append(['Fractal EMG1'])
r = 1
c = 1
cellcheck = 0
for i in range(len(emg1filt)//750):
    emg_sec = emg1filt[0+(750*i):7500+(750*i)]
    frac1 = _higuchi_fd(emg_sec,6)
    while cellcheck == 0:
        if sheet.cell(row=r,column=c).value != None:
            r = r+1
        else:
            cellcheck = 1
    sheet.cell(row=r, column=c).value = frac1
    r = r + 1
    # print(frac1)

book.save(filename=dest_filename)


#### add an option to put multiple files inputted and multiple emgs captured all outputted to 1 excel file
### do 1) mean frequency, 2) median frequency, 3) power amp, 4) spectral moments, 5) timing, 6) fractals
#### DO ALL ABOVE IN SEPARATE PYTHON FILE! USE THIS FILE TO CHECK APPEND OF

##  url for better excel handling = https://stackoverflowcom/questions/202119254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
## look at second answer