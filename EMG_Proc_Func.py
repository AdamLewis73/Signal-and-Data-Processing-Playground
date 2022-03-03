from scipy import fftpack, signal
from matplotlib import pyplot as plt
import math
import numpy as np
import pandas as pd
import csv
import scipy.signal as sps
from scipy.signal import hilbert

from math import floor,log

temp = []
samplerate = 1500
nyq = samplerate*0.5

def _linear_regression(x, y):
    """Fast linear regression using Numba.
    Parameters
    ----------
    x, y : ndarray, shape (n_times,)
        Variables
    Returns
    -------
    slope : float
        Slope of 1D least-square regression.
    intercept : float
        Intercept
    """
    n_times = x.size
    sx2 = 0
    sx = 0
    sy = 0
    sxy = 0
    for j in range(n_times):
        sx2 += x[j] ** 2
        sx += x[j]
        sxy += x[j] * y[j]
        sy += y[j]
    den = n_times * sx2 - (sx ** 2)
    num = n_times * sxy - sx * sy
    slope = num / den
    intercept = np.mean(y) - slope * np.mean(x)
    return slope, intercept

def _higuchi_fd(x, kmax):
    n_times = x.size
    lk = np.empty(kmax)
    x_reg = np.empty(kmax)
    y_reg = np.empty(kmax)
    for k in range(1, kmax + 1):
        lm = np.empty((k,))
        for m in range(k):
            ll = 0
            n_max = floor((n_times - m - 1) / k)
            n_max = int(n_max)
            for j in range(1, n_max):
                ll += abs(x[m + j * k] - x[m + (j - 1) * k])
            ll /= k
            ll *= (n_times - 1) / (k * n_max)
            lm[m] = ll
        # Mean of lm
        m_lm = 0
        for m in range(k):
            m_lm += lm[m]
        m_lm /= k
        lk[k - 1] = m_lm
        x_reg[k - 1] = log(1. / k)
        y_reg[k - 1] = log(m_lm)
    higuchi, _ = _linear_regression(x_reg, y_reg)
    return higuchi


from openpyxl import Workbook
from openpyxl.chart import (ScatterChart,Reference,Series)
import openpyxl


dest_filename = 'Subj08.xlsx'
book = Workbook()
# book = openpyxl.load_workbook(dest_filename)
sheet = book.active

filename = 'C:/Users/sword/Anaconda3/envs/exceltest/RF_Subj08/Free/RF_Subj08_Free_Fatigue5_TR01.csv'
# filename = 'C:/Users/sword/Anaconda3/envs/exceltest/RF_Subj08/Free/RF_Subj08_Free_Fatigue5_TR24.csv'
import openpyxl

with open(filename, 'r') as csvfile:
    csvreader = csv.reader(csvfile, delimiter=',')
    for row in csvreader:
        if csvreader.line_num == 3:
            temp.append(row)
        if csvreader.line_num >= 6:
            if row:
                temp.append(row)
            else:
                break

df = pd.DataFrame(temp)  # turns the array into a dataframe
df.columns = df.iloc[0]  # sets the column names as the first row
df = df.drop(0)  # drops the first row since it is now a duplicate of the column names
df.reindex(df.index.drop(1))
df.reset_index(drop=True, inplace=True)
df.columns = ['frames', 'subframes', 'blank', 'emg1', 'emg2', 'emg3', 'emg4', 'emg5', 'emg6', 'emg7', 'emg8', 'emg9', 'emg10', 'emg11', 'emg12', 'emg13', 'emg14', 'emg15', 'emg16', 'unused7', 'unused8', 'blank2']
df2 = df.drop(['frames', 'subframes', 'blank', 'unused7', 'unused8', 'blank2'], axis=1)
df2 = df2.astype(np.float)
print(len(df))
hor = np.arange(0, (len(df) - 0.5) / samplerate, 1 / samplerate)  # getting the time domain in seconds

emg1 = df2.emg2
emg2 = df2.emg4
emg3 = df2.emg6
emg4 = df2.emg8
emg5 = df2.emg10
emg6 = df2.emg6
emg7 = df2.emg8

cutoff_freq = 20  # ~20 Hz for movement according to emg book "Electromyography: Physiology, Engineering, and Noninvasive Apps"
cut = cutoff_freq/nyq
b, a = signal.butter(5, cut, btype='highpass', analog=False)
emg1high = signal.filtfilt(b, a, emg1)
emg2high = signal.filtfilt(b, a, emg2)
emg3high = signal.filtfilt(b, a, emg3)
emg4high = signal.filtfilt(b, a, emg4)
emg5high = signal.filtfilt(b, a, emg5)
emg6high = signal.filtfilt(b, a, emg6)
emg7high = signal.filtfilt(b, a, emg7)

cutoff_freq = 400  # ~500 Hz according to the emg book
cut = cutoff_freq/nyq
b, a = signal.butter(5, cut, btype='lowpass', analog=False)
emg1filt = signal.filtfilt(b, a, emg1high)
emg2filt = signal.filtfilt(b, a, emg2high)
emg3filt = signal.filtfilt(b, a, emg3high)
emg4filt = signal.filtfilt(b, a, emg4high)
emg5filt = signal.filtfilt(b, a, emg5high)
emg6filt = signal.filtfilt(b, a, emg6high)
emg7filt = signal.filtfilt(b, a, emg7high)

plt.figure(1)
plt.subplot(2,4,1)
plt.plot(hor,emg1)
plt.subplot(2,4,2)
plt.plot(hor,emg2)
plt.subplot(2,4,3)
plt.plot(hor,emg3)
plt.subplot(2,4,4)
plt.plot(hor,emg4)
plt.subplot(2,4,5)
plt.plot(hor,emg5)
plt.subplot(2,4,6)
plt.plot(hor,emg6)
plt.subplot(2,4,7)
plt.plot(hor,emg7)

# plt.show()

emg_rec1 = abs(emg1filt)
emg_rec2 = abs(emg2filt)
emg_rec3 = abs(emg3filt)
emg_rec4 = abs(emg4filt)
emg_rec5 = abs(emg5filt)
emg_rec6 = abs(emg6filt)
emg_rec7 = abs(emg7filt)
ynew1 = signal.savgol_filter(emg_rec1,1501,2)
ynew2 = signal.savgol_filter(emg_rec2,1501,2)
ynew3 = signal.savgol_filter(emg_rec3,1501,2)
ynew4 = signal.savgol_filter(emg_rec4,1501,2)
ynew5 = signal.savgol_filter(emg_rec5,1501,2)
ynew6 = signal.savgol_filter(emg_rec6,1501,2)
ynew7 = signal.savgol_filter(emg_rec7,1501,2)

cutoff_freq = 10  # ~20 Hz for movement according to emg book "Electromyography: Physiology, Engineering, and Noninvasive Apps"
cut = cutoff_freq/nyq
b, a = signal.butter(5, cut, btype='lowpass', analog=False)

ynew1 = signal.filtfilt(b, a, ynew1)
ynew2 = signal.filtfilt(b, a, ynew2)
ynew3 = signal.filtfilt(b, a, ynew3)
ynew4 = signal.filtfilt(b, a, ynew4)
ynew5 = signal.filtfilt(b, a, ynew5)
ynew6 = signal.filtfilt(b, a, ynew6)
ynew7 = signal.filtfilt(b, a, ynew7)

ynew1 = signal.savgol_filter(ynew1,1501,2)
ynew2 = signal.savgol_filter(ynew2,1501,2)
ynew3 = signal.savgol_filter(ynew3,1501,2)
ynew4 = signal.savgol_filter(ynew4,1501,2)
ynew5 = signal.savgol_filter(ynew5,1501,2)
ynew6 = signal.savgol_filter(ynew6,1501,2)
ynew7 = signal.savgol_filter(ynew7,1501,2)

peaks1, props1 = signal.find_peaks(ynew1, width=(500,2000), rel_height=0.8)
peaks2, props2 = signal.find_peaks(ynew2, width=(500,2000), rel_height=0.8)
peaks3, props3 = signal.find_peaks(ynew3, width=(500,2000), rel_height=0.8)
peaks4, props4 = signal.find_peaks(ynew4, width=(500,2000), rel_height=0.8)
peaks5, props5 = signal.find_peaks(ynew5, width=(500,2000), rel_height=0.8)
peaks6, props6 = signal.find_peaks(ynew6, width=(500,2000), rel_height=0.8)
peaks7, props7 = signal.find_peaks(ynew7, width=(500,2000), rel_height=0.8)
print("len of peaks1: " + str(len(peaks1)))

pulses_beginT1 = []
pulses_beginT2 = []
pulses_beginT3 = []
pulses_beginT4 = []
pulses_beginT5 = []
pulses_beginT6 = []
pulses_beginT7 = []
pulses_endT1 = []
pulses_endT2 = []
pulses_endT3 = []
pulses_endT4 = []
pulses_endT5 = []
pulses_endT6 = []
pulses_endT7 = []
pulses_begin1 = []
pulses_begin2 = []
pulses_begin3 = []
pulses_begin4 = []
pulses_begin5 = []
pulses_begin6 = []
pulses_begin7 = []
pulses_end1 = []
pulses_end2 = []
pulses_end3 = []
pulses_end4 = []
pulses_end5 = []
pulses_end6 = []
pulses_end7 = []
pulses_begin_ind1 = []
pulses_begin_ind2 = []
pulses_begin_ind3 = []
pulses_begin_ind4 = []
pulses_begin_ind5 = []
pulses_begin_ind6 = []
pulses_begin_ind7 = []
pulses_end_ind1 = []
pulses_end_ind2 = []
pulses_end_ind3 = []
pulses_end_ind4 = []
pulses_end_ind5 = []
pulses_end_ind6 = []
pulses_end_ind7 = []

for i in range(len(peaks1)):
    pulse_sample_start1 = peaks1[i] - (math.floor(props1['widths'][i]/2))
    pulse_sample_end1 = peaks1[i] + (math.floor(props1['widths'][i]/2))
    pulses_begin_ind1.append(pulse_sample_start1)
    pulses_end_ind1.append(pulse_sample_end1)
    pulses_beginT1.append(pulse_sample_start1/1500)
    pulses_endT1.append(pulse_sample_end1/1500)
    pulses_begin1.append(ynew1[pulse_sample_start1])
    pulses_end1.append(ynew1[pulse_sample_end1])

for i in range(len(peaks2)):
    pulse_sample_start2 = peaks2[i] - (math.floor(props2['widths'][i]/2))
    pulse_sample_end2 = peaks2[i] + (math.floor(props2['widths'][i]/2))
    pulses_begin_ind2.append(pulse_sample_start2)
    pulses_end_ind2.append(pulse_sample_end2)
    pulses_beginT2.append(pulse_sample_start2/1500)
    pulses_endT2.append(pulse_sample_end2/1500)
    pulses_begin2.append(ynew2[pulse_sample_start2])
    pulses_end2.append(ynew2[pulse_sample_end2])

for i in range(len(peaks3)):
    pulse_sample_start3 = peaks3[i] - (math.floor(props3['widths'][i]/2))
    pulse_sample_end3 = peaks3[i] + (math.floor(props3['widths'][i]/2))
    pulses_begin_ind3.append(pulse_sample_start3)
    pulses_end_ind3.append(pulse_sample_end3)
    pulses_beginT3.append(pulse_sample_start3/1500)
    pulses_endT3.append(pulse_sample_end3/1500)
    pulses_begin3.append(ynew3[pulse_sample_start3])
    pulses_end3.append(ynew3[pulse_sample_end3])

for i in range(len(peaks4)):
    pulse_sample_start4 = peaks4[i] - (math.floor(props4['widths'][i]/2))
    pulse_sample_end4 = peaks4[i] + (math.floor(props4['widths'][i]/2))
    pulses_begin_ind4.append(pulse_sample_start4)
    pulses_end_ind4.append(pulse_sample_end4)
    pulses_beginT4.append(pulse_sample_start4/1500)
    pulses_endT4.append(pulse_sample_end4/1500)
    pulses_begin4.append(ynew4[pulse_sample_start4])
    pulses_end4.append(ynew4[pulse_sample_end4])

for i in range(len(peaks5)):
    pulse_sample_start5 = peaks5[i] - (math.floor(props5['widths'][i]/2))
    pulse_sample_end5 = peaks5[i] + (math.floor(props5['widths'][i]/2))
    pulses_begin_ind5.append(pulse_sample_start5)
    pulses_end_ind5.append(pulse_sample_end5)
    pulses_beginT5.append(pulse_sample_start5/1500)
    pulses_endT5.append(pulse_sample_end5/1500)
    pulses_begin5.append(ynew5[pulse_sample_start5])
    pulses_end5.append(ynew5[pulse_sample_end5])

for i in range(len(peaks6)):
    pulse_sample_start6 = peaks6[i] - (math.floor(props6['widths'][i]/2))
    pulse_sample_end6 = peaks6[i] + (math.floor(props6['widths'][i]/2))
    pulses_begin_ind6.append(pulse_sample_start6)
    pulses_end_ind6.append(pulse_sample_end6)
    pulses_beginT6.append(pulse_sample_start6/1500)
    pulses_endT6.append(pulse_sample_end6/1500)
    pulses_begin6.append(ynew6[pulse_sample_start6])
    pulses_end6.append(ynew6[pulse_sample_end6])

for i in range(len(peaks7)):
    pulse_sample_start7 = peaks7[i] - (math.floor(props7['widths'][i]/2))
    pulse_sample_end7 = peaks7[i] + (math.floor(props7['widths'][i]/2))
    pulses_begin_ind7.append(pulse_sample_start7)
    pulses_end_ind7.append(pulse_sample_end7)
    pulses_beginT7.append(pulse_sample_start7/1500)
    pulses_endT7.append(pulse_sample_end7/1500)
    pulses_begin7.append(ynew7[pulse_sample_start7])
    pulses_end7.append(ynew7[pulse_sample_end7])

M = 0
N = 1
lowcut = 10
highcut = 400
cellcheck = 0
for a in range((len(peaks1))):

    sectionpoints1 = []
    sectionT1 = []
    for i in range(M, N):
        sectionpoints1.extend(emg_rec1[pulses_begin_ind1[i]:pulses_end_ind1[i]])
        sectionT1.extend(hor[pulses_begin_ind1[i]:pulses_end_ind1[i]])

    x1 = np.linspace(0, len(sectionpoints1), len(sectionpoints1))
    x1 = x1 / 1500

    sectionpoints_array1 = np.asarray(sectionpoints1)
    freq1, power_spec1 = signal.periodogram(sectionpoints_array1, samplerate)

    lowflag = 0
    for c in range(len(freq1)):
        if abs(freq1[c]) > 10 and lowflag == 0:
            fullpulse_lowfreqbound = c
            lowflag = 1
        if abs(freq1[c]) > 400:
            fullpulse_uppfreqbound = c
            break

    spec_mom0 = 0
    spec_mom2 = 0
    spec_mom5 = 0

    for i in range(fullpulse_lowfreqbound, fullpulse_uppfreqbound):
        spec_mom0 = spec_mom0 + (math.pow(freq1[i], -1) * power_spec1[i])
        spec_mom2 = spec_mom2 + (math.pow(freq1[i], 2) * power_spec1[i])
        spec_mom5 = spec_mom5 + (math.pow(freq1[i], 5) * power_spec1[i])

    f2 = spec_mom0 / spec_mom2
    f5 = spec_mom0 / spec_mom5

    powsum = 0
    powarray = []

    for i in range(len(freq1)):
        powsum = powsum + power_spec1[i]
        powarray.append(powsum)

    mednum = powsum / 2

    meansumcombo = 0
    meansumpow = 0

    for i in range(len(freq1)):
        meansumcombo = meansumcombo + (freq1[i]*power_spec1[i])
        meansumpow = meansumpow + (power_spec1[i])

    mean = meansumcombo / meansumpow

    row = 2
    col = 1
    cellcheck = 0
    for i in powarray:
        if i > mednum:
            median = freq1[powarray.index(i)]
            while cellcheck == 0:
                if sheet.cell(row=row, column=col).value != None:
                    col = col + 1
                else:
                    cellcheck = 1
            sheet.cell(row=row, column=col).value = median
            break

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = mean

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f2

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f5

    M = M + 1
    N = N + 1


row = 10
col = 1
cellcheck = 0
for i in range(len(emg1filt) // 750):
    emg_sec = emg1filt[(0 + (750 * i)):(7500 + (750 * i))]
    frac1 = _higuchi_fd(emg_sec, 6)
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = frac1
    col = col+1
# ^row 2, 4, 6, 8, 10, 12

M = 0
N = 1
cellcheck = 0
for a in range((len(peaks2))):

    sectionpoints2 = []
    sectionT2 = []
    for i in range(M, N):
        sectionpoints2.extend(emg_rec2[pulses_begin_ind2[i]:pulses_end_ind2[i]])
        sectionT2.extend(hor[pulses_begin_ind2[i]:pulses_end_ind2[i]])

    x2 = np.linspace(0, len(sectionpoints2), len(sectionpoints2))
    x2 = x2 / 1500

    sectionpoints_array2 = np.asarray(sectionpoints2)
    freq2, power_spec2 = signal.periodogram(sectionpoints_array2, samplerate)

    lowflag = 0
    for c in range(len(freq2)):
        if abs(freq2[c]) > 10 and lowflag == 0:
            fullpulse_lowfreqbound = c
            lowflag = 1
        if abs(freq2[c]) > 400:
            fullpulse_uppfreqbound = c
            break

    spec_mom0 = 0
    spec_mom2 = 0
    spec_mom5 = 0

    for i in range(fullpulse_lowfreqbound, fullpulse_uppfreqbound):
        spec_mom0 = spec_mom0 + (math.pow(freq2[i], -1) * power_spec2[i])
        spec_mom2 = spec_mom2 + (math.pow(freq2[i], 2) * power_spec2[i])
        spec_mom5 = spec_mom5 + (math.pow(freq2[i], 5) * power_spec2[i])

    f2 = spec_mom0 / spec_mom2
    f5 = spec_mom0 / spec_mom5

    powsum = 0
    powarray = []

    for i in range(len(freq2)):
        powsum = powsum + power_spec2[i]
        powarray.append(powsum)

    mednum = powsum / 2

    meansumcombo = 0
    meansumpow = 0

    for i in range(len(freq2)):
        meansumcombo = meansumcombo + (freq2[i] * power_spec2[i])
        meansumpow = meansumpow + (power_spec2[i])

    mean = meansumcombo / meansumpow

    row = 17
    col = 1
    cellcheck = 0
    for i in powarray:
        if i > mednum:
            median = freq2[powarray.index(i)]
            while cellcheck == 0:
                if sheet.cell(row=row, column=col).value != None:
                    col = col + 1
                else:
                    cellcheck = 1
            sheet.cell(row=row, column=col).value = median
            break

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = mean

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f2

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f5

    M = M + 1
    N = N + 1


row = 25
col = 1
cellcheck = 0
for i in range(len(emg2filt) // 750):
    emg_sec = emg2filt[(0 + (750 * i)):(7500 + (750 * i))]
    frac1 = _higuchi_fd(emg_sec, 6)
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = frac1
    col = col + 1

# ^row 17, 19, 21, 23, 25, 27

M = 0
N = 1
cellcheck = 0
for a in range((len(peaks3))):

    sectionpoints3 = []
    sectionT3 = []
    for i in range(M, N):
        sectionpoints3.extend(emg_rec3[pulses_begin_ind3[i]:pulses_end_ind3[i]])
        sectionT3.extend(hor[pulses_begin_ind3[i]:pulses_end_ind3[i]])

    x3 = np.linspace(0, len(sectionpoints3), len(sectionpoints3))
    x3 = x3 / 1500

    sectionpoints_array3 = np.asarray(sectionpoints3)
    freq3, power_spec3 = signal.periodogram(sectionpoints_array3, samplerate)

    lowflag = 0
    for c in range(len(freq3)):
        if abs(freq3[c]) > 10 and lowflag == 0:
            fullpulse_lowfreqbound = c
            lowflag = 1
        if abs(freq3[c]) > 400:
            fullpulse_uppfreqbound = c
            break

    spec_mom0 = 0
    spec_mom2 = 0
    spec_mom5 = 0

    for i in range(fullpulse_lowfreqbound, fullpulse_uppfreqbound):
        spec_mom0 = spec_mom0 + (math.pow(freq3[i], -1) * power_spec3[i])
        spec_mom2 = spec_mom2 + (math.pow(freq3[i], 2) * power_spec3[i])
        spec_mom5 = spec_mom5 + (math.pow(freq3[i], 5) * power_spec3[i])

    f2 = spec_mom0 / spec_mom2
    f5 = spec_mom0 / spec_mom5

    powsum = 0
    powarray = []

    for i in range(len(freq3)):
        powsum = powsum + power_spec3[i]
        powarray.append(powsum)

    mednum = powsum / 2

    meansumcombo = 0
    meansumpow = 0

    for i in range(len(freq3)):
        meansumcombo = meansumcombo + (freq3[i] * power_spec3[i])
        meansumpow = meansumpow + (power_spec3[i])

    mean = meansumcombo / meansumpow

    row = 32
    col = 1
    cellcheck = 0
    for i in powarray:
        if i > mednum:
            median = freq3[powarray.index(i)]
            while cellcheck == 0:
                if sheet.cell(row=row, column=col).value != None:
                    col = col + 1
                else:
                    cellcheck = 1
            sheet.cell(row=row, column=col).value = median
            break

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = mean

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f2

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f5

    M = M + 1
    N = N + 1


row = 40
col = 1
cellcheck = 0
for i in range(len(emg3filt) // 750):
    emg_sec = emg3filt[(0 + (750 * i)):(7500 + (750 * i))]
    frac1 = _higuchi_fd(emg_sec, 6)
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = frac1
    col = col + 1

# ^row 32, 34, 36, 38, 40, 42

M = 0
N = 1
cellcheck = 0
for a in range((len(peaks4))):

    sectionpoints4 = []
    sectionT4 = []
    for i in range(M, N):
        sectionpoints4.extend(emg_rec4[pulses_begin_ind4[i]:pulses_end_ind4[i]])
        sectionT4.extend(hor[pulses_begin_ind4[i]:pulses_end_ind4[i]])

    x4 = np.linspace(0, len(sectionpoints4), len(sectionpoints4))
    x4 = x4 / 1500

    sectionpoints_array4 = np.asarray(sectionpoints4)
    freq4, power_spec4 = signal.periodogram(sectionpoints_array4, samplerate)

    lowflag = 0
    for c in range(len(freq4)):
        if abs(freq4[c]) > 10 and lowflag == 0:
            fullpulse_lowfreqbound = c
            lowflag = 1
        if abs(freq4[c]) > 400:
            fullpulse_uppfreqbound = c
            break

    spec_mom0 = 0
    spec_mom2 = 0
    spec_mom5 = 0

    for i in range(fullpulse_lowfreqbound, fullpulse_uppfreqbound):
        spec_mom0 = spec_mom0 + (math.pow(freq4[i], -1) * power_spec4[i])
        spec_mom2 = spec_mom2 + (math.pow(freq4[i], 2) * power_spec4[i])
        spec_mom5 = spec_mom5 + (math.pow(freq4[i], 5) * power_spec4[i])

    f2 = spec_mom0 / spec_mom2
    f5 = spec_mom0 / spec_mom5

    powsum = 0
    powarray = []

    for i in range(len(freq4)):
        powsum = powsum + power_spec4[i]
        powarray.append(powsum)

    mednum = powsum / 2

    meansumcombo = 0
    meansumpow = 0

    for i in range(len(freq4)):
        meansumcombo = meansumcombo + (freq4[i] * power_spec4[i])
        meansumpow = meansumpow + (power_spec4[i])

    mean = meansumcombo / meansumpow

    row = 47
    col = 1
    cellcheck = 0
    for i in powarray:
        if i > mednum:
            median = freq4[powarray.index(i)]
            while cellcheck == 0:
                if sheet.cell(row=row, column=col).value != None:
                    col = col + 1
                else:
                    cellcheck = 1
            sheet.cell(row=row, column=col).value = median
            break

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = mean

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f2

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f5

    M = M + 1
    N = N + 1


row = 55
col = 1
cellcheck = 0
for i in range(len(emg4filt) // 750):
    emg_sec = emg4filt[(0 + (750 * i)):(7500 + (750 * i))]
    frac1 = _higuchi_fd(emg_sec, 6)
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = frac1
    col = col + 1

# ^row 47, 49, 51, 53, 55, 57

M = 0
N = 1
cellcheck = 0
for a in range((len(peaks5))):

    sectionpoints5 = []
    sectionT5 = []
    for i in range(M, N):
        sectionpoints5.extend(emg_rec5[pulses_begin_ind5[i]:pulses_end_ind5[i]])
        sectionT5.extend(hor[pulses_begin_ind5[i]:pulses_end_ind5[i]])

    x5 = np.linspace(0, len(sectionpoints5), len(sectionpoints5))
    x5 = x5 / 1500

    sectionpoints_array5 = np.asarray(sectionpoints5)
    freq5, power_spec5 = signal.periodogram(sectionpoints_array5, samplerate)

    lowflag = 0
    for c in range(len(freq5)):
        if abs(freq5[c]) > 10 and lowflag == 0:
            fullpulse_lowfreqbound = c
            lowflag = 1
        if abs(freq5[c]) > 400:
            fullpulse_uppfreqbound = c
            break

    spec_mom0 = 0
    spec_mom2 = 0
    spec_mom5 = 0

    for i in range(fullpulse_lowfreqbound, fullpulse_uppfreqbound):
        spec_mom0 = spec_mom0 + (math.pow(freq5[i], -1) * power_spec5[i])
        spec_mom2 = spec_mom2 + (math.pow(freq5[i], 2) * power_spec5[i])
        spec_mom5 = spec_mom5 + (math.pow(freq5[i], 5) * power_spec5[i])

    f2 = spec_mom0 / spec_mom2
    f5 = spec_mom0 / spec_mom5

    powsum = 0
    powarray = []

    for i in range(len(freq5)):
        powsum = powsum + power_spec5[i]
        powarray.append(powsum)

    mednum = powsum / 2

    meansumcombo = 0
    meansumpow = 0

    for i in range(len(freq5)):
        meansumcombo = meansumcombo + (freq5[i] * power_spec5[i])
        meansumpow = meansumpow + (power_spec5[i])

    mean = meansumcombo / meansumpow

    row = 47
    col = 1
    cellcheck = 0
    for i in powarray:
        if i > mednum:
            median = freq5[powarray.index(i)]
            while cellcheck == 0:
                if sheet.cell(row=row, column=col).value != None:
                    col = col + 1
                else:
                    cellcheck = 1
            sheet.cell(row=row, column=col).value = median
            break

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = mean

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f2

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f5

    M = M + 1
    N = N + 1


row = 70
col = 1
cellcheck = 0
for i in range(len(emg5filt) // 750):
    emg_sec = emg5filt[(0 + (750 * i)):(7500 + (750 * i))]
    frac1 = _higuchi_fd(emg_sec, 6)
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = frac1
    col = col + 1
# ^row 62, 64, 66, 68, 70, 72

M = 0
N = 1
cellcheck = 0
for a in range((len(peaks6))):

    sectionpoints6 = []
    sectionT6 = []
    for i in range(M, N):
        sectionpoints6.extend(emg_rec6[pulses_begin_ind6[i]:pulses_end_ind6[i]])
        sectionT6.extend(hor[pulses_begin_ind6[i]:pulses_end_ind6[i]])

    x6 = np.linspace(0, len(sectionpoints6), len(sectionpoints6))
    x6 = x6 / 1500

    sectionpoints_array6 = np.asarray(sectionpoints6)
    freq6, power_spec6 = signal.periodogram(sectionpoints_array6, samplerate)

    lowflag = 0
    for c in range(len(freq6)):
        if abs(freq6[c]) > 10 and lowflag == 0:
            fullpulse_lowfreqbound = c
            lowflag = 1
        if abs(freq6[c]) > 400:
            fullpulse_uppfreqbound = c
            break

    spec_mom0 = 0
    spec_mom2 = 0
    spec_mom5 = 0

    for i in range(fullpulse_lowfreqbound, fullpulse_uppfreqbound):
        spec_mom0 = spec_mom0 + (math.pow(freq6[i], -1) * power_spec6[i])
        spec_mom2 = spec_mom2 + (math.pow(freq6[i], 2) * power_spec6[i])
        spec_mom5 = spec_mom5 + (math.pow(freq6[i], 5) * power_spec6[i])

    f2 = spec_mom0 / spec_mom2
    f5 = spec_mom0 / spec_mom5

    powsum = 0
    powarray = []

    for i in range(len(freq6)):
        powsum = powsum + power_spec6[i]
        powarray.append(powsum)

    mednum = powsum / 2

    meansumcombo = 0
    meansumpow = 0

    for i in range(len(freq6)):
        meansumcombo = meansumcombo + (freq6[i] * power_spec6[i])
        meansumpow = meansumpow + (power_spec6[i])

    mean = meansumcombo / meansumpow

    row = 62
    col = 1
    cellcheck = 0
    for i in powarray:
        if i > mednum:
            median = freq6[powarray.index(i)]
            while cellcheck == 0:
                if sheet.cell(row=row, column=col).value != None:
                    col = col + 1
                else:
                    cellcheck = 1
            sheet.cell(row=row, column=col).value = median
            break

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = mean

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f2

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f5

    M = M + 1
    N = N + 1


row = 85
col = 1
cellcheck = 0
for i in range(len(emg6filt) // 750):
    emg_sec = emg6filt[(0 + (750 * i)):(7500 + (750 * i))]
    frac1 = _higuchi_fd(emg_sec, 6)
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = frac1
    col = col + 1

# ^row 77, 79, 81, 83, 85, 87

M = 0
N = 1
cellcheck = 0
for a in range((len(peaks7))):

    sectionpoints7 = []
    sectionT7 = []
    for i in range(M, N):
        sectionpoints7.extend(emg_rec7[pulses_begin_ind7[i]:pulses_end_ind7[i]])
        sectionT7.extend(hor[pulses_begin_ind7[i]:pulses_end_ind7[i]])

    x7 = np.linspace(0, len(sectionpoints7), len(sectionpoints7))
    x7 = x7 / 1500

    sectionpoints_array7 = np.asarray(sectionpoints7)
    freq7, power_spec7 = signal.periodogram(sectionpoints_array7, samplerate)

    lowflag = 0
    for c in range(len(freq7)):
        if abs(freq7[c]) > 10 and lowflag == 0:
            fullpulse_lowfreqbound = c
            lowflag = 1
        if abs(freq7[c]) > 400:
            fullpulse_uppfreqbound = c
            break

    spec_mom0 = 0
    spec_mom2 = 0
    spec_mom5 = 0
    # print(len(freq7))
    # print(len(power_spec7))
    for i in range(fullpulse_lowfreqbound, fullpulse_uppfreqbound):
        spec_mom0 = spec_mom0 + (math.pow(freq7[i], -1) * power_spec7[i])
        spec_mom2 = spec_mom2 + (math.pow(freq7[i], 2) * power_spec7[i])
        spec_mom5 = spec_mom5 + (math.pow(freq7[i], 5) * power_spec7[i])

    f2 = spec_mom0 / spec_mom2
    f5 = spec_mom0 / spec_mom5

    powsum = 0
    powarray = []

    for i in range(len(freq7)):
        powsum = powsum + power_spec7[i]
        powarray.append(powsum)

    mednum = powsum / 2

    meansumcombo = 0
    meansumpow = 0

    for i in range(len(freq7)):
        meansumcombo = meansumcombo + (freq7[i] * power_spec7[i])
        meansumpow = meansumpow + (power_spec7[i])

    mean = meansumcombo / meansumpow

    row = 77
    col = 1
    cellcheck = 0
    for i in powarray:
        if i > mednum:
            median = freq7[powarray.index(i)]
            while cellcheck == 0:
                if sheet.cell(row=row, column=col).value != None:
                    col = col + 1
                else:
                    cellcheck = 1
            sheet.cell(row=row, column=col).value = median
            break

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = mean

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f2

    cellcheck = 0
    row = row + 2
    col = 1
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = f5

    M = M + 1
    N = N + 1


row = 100
col = 1
cellcheck = 0
for i in range(len(emg7filt) // 750):
    emg_sec = emg7filt[(0 + (750 * i)):(7500 + (750 * i))]
    frac1 = _higuchi_fd(emg_sec, 6)
    while cellcheck == 0:
        if sheet.cell(row=row, column=col).value != None:
            col = col + 1
        else:
            cellcheck = 1
    sheet.cell(row=row, column=col).value = frac1
    col = col + 1

# ^row 92, 94, 96, 98, 100, 102


# chart = ScatterChart()
# chart.title = "EMG 1 Median Frequency"
# chart.style = 13
# chart.x_axis.title = 'Pulse'
# chart.y_axis.title = 'Frequency'
#
# count = 0
# while sheet.cell(row=2, column=count+1).value != None:
#     count = count+1
# xval = list(range(count))
# yval = Reference(sheet, min_col = 1, min_row = 2, max_col = count, max_row =2)
# series = Series(values=yval, xvalues=xval)
#
# chart.series.append(series)
#
# sheet.add_chart(chart, 'A10')

# chart.title = "EMG 1 Mean Frequency"
#
# chart.title = "EMG 1 Spectral Moment Order 2"
#
# chart.title = "EMG 1 Spectral Moment Order 5"
#
# chart.title = "EMG 1 Fractal Dimension"




book.save(filename=dest_filename)


fig = plt.figure(2)
ax0 = fig.add_subplot(2, 4, 1)
ax0.plot(hor, emg_rec1)
ax0.plot(hor,ynew1,color='red')
ax0.plot(pulses_beginT1, pulses_begin1, marker = "x",color = 'yellow', linestyle = "None")
ax0.plot(pulses_endT1, pulses_end1, marker = "x",color = 'yellow', linestyle = "None")

ax0 = fig.add_subplot(2, 4, 2)
ax0.plot(hor, emg_rec2)
ax0.plot(hor,ynew2,color='red')
ax0.plot(pulses_beginT2, pulses_begin2, marker = "x",color = 'yellow', linestyle = "None")
ax0.plot(pulses_endT2, pulses_end2, marker = "x",color = 'yellow', linestyle = "None")

ax0 = fig.add_subplot(2, 4, 3)
ax0.plot(hor, emg_rec3)
ax0.plot(hor,ynew3,color='red')
ax0.plot(pulses_beginT3, pulses_begin3, marker = "x",color = 'yellow', linestyle = "None")
ax0.plot(pulses_endT3, pulses_end3, marker = "x",color = 'yellow', linestyle = "None")

ax0 = fig.add_subplot(2, 4, 4)
ax0.plot(hor, emg_rec4)
ax0.plot(hor,ynew4,color='red')
ax0.plot(pulses_beginT4, pulses_begin4, marker = "x",color = 'yellow', linestyle = "None")
ax0.plot(pulses_endT4, pulses_end4, marker = "x",color = 'yellow', linestyle = "None")

ax0 = fig.add_subplot(2, 4, 5)
ax0.plot(hor, emg_rec5)
ax0.plot(hor,ynew5,color='red')
ax0.plot(pulses_beginT5, pulses_begin5, marker = "x",color = 'yellow', linestyle = "None")
ax0.plot(pulses_endT5, pulses_end5, marker = "x",color = 'yellow', linestyle = "None")

ax0 = fig.add_subplot(2, 4, 6)
ax0.plot(hor, emg_rec6)
ax0.plot(hor,ynew6,color='red')
ax0.plot(pulses_beginT6, pulses_begin6, marker = "x",color = 'yellow', linestyle = "None")
ax0.plot(pulses_endT6, pulses_end6, marker = "x",color = 'yellow', linestyle = "None")

ax0 = fig.add_subplot(2, 4, 7)
ax0.plot(hor, emg_rec7)
ax0.plot(hor,ynew7,color='red')
ax0.plot(pulses_beginT7, pulses_begin7, marker = "x",color = 'yellow', linestyle = "None")
ax0.plot(pulses_endT7, pulses_end7, marker = "x",color = 'yellow', linestyle = "None")

# plt.show()