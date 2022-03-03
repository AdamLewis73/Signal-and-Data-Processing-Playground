from tkinter import *
from tkinter import filedialog
# from tkMessageBox import *
import tkinter

window = Tk()
menu = Menu(window)
d= {}
# for i in enumerate(range(1,17)):
#     d['v{}'.format(i)] = tkinter.IntVar()
for i in range(1,17):
    d['v{}'.format(i)] = tkinter.IntVar()
# for i in range(1,17):
#     print(d['v{}'.format(i)])
# exit()
# v[1] = tkinter.IntVar()
# v[2] = tkinter.IntVar()
# v[3] = tkinter.IntVar()
# v[4] = tkinter.IntVar()
# v[5] = tkinter.IntVar()
# v[6] = tkinter.IntVar()
# v[7] = tkinter.IntVar()
# v[8] = tkinter.IntVar()
# v[9] = tkinter.IntVar()
# v[10] = tkinter.IntVar()
# v[11] = tkinter.IntVar()
# v[12] = tkinter.IntVar()
# v[13] = tkinter.IntVar()
# v[14] = tkinter.IntVar()
# v[15] = tkinter.IntVar()
# v[16] = tkinter.IntVar()



emg_options = [('EMG 1',1),('EMG 2',2),('EMG 3',3),('EMG 4',4),
               ('EMG 5',5),('EMG 6',6),('EMG 7',7),('EMG 8',8),
               ('EMG 9',9),('EMG 10',10),('EMG 11',11),('EMG 12',12),
               ('EMG 13',13),('EMG 14',14),('EMG 15',15),('EMG 16',16)]

# file = filedialog.askopenfilename()
filename = ''
window.title('EMGProc')
# window.geometry('800x600')
csv_checker = 0
color_iter = 0
filenamesplit =[]
emg_choice = []
overwrite_choice = 0
relative_height = 0.8
pulse_width_min = 500
pulse_width_max = 2000
k_val = 6
# import EMG_Proc_Func.py


def change_color():
    global color_iter
    global bg_color
    current_color = file_txt.cget("background")
    next_color = bg_color if current_color == "red" else "red"
    file_txt.config(background=next_color)
    if color_iter < 4:
        window.after(500, change_color)
        color_iter += 1
    else:
        color_iter = 0
        return

def dataproc():
    global csv_checker
    global filename
    global filenamesplit
    global emg_choice
    global overwrite_choice
    global output_filename_choice
    global relative_height
    global pulse_width_min
    global pulse_width_max
    global k_val
    # print(len(filename))
    # print(filename)
    # exit()

    if csv_checker == 0:
        from scipy import fftpack, signal
        from matplotlib import pyplot as plt
        import math
        import numpy as np
        import pandas as pd
        import csv

        import scipy.signal as sps
        from scipy.signal import hilbert

        from math import floor, log

        # temp = []
        samplerate = 1500
        nyq = samplerate * 0.5
        row_buff_val2 = 15
        kmin = 2

        def _linear_regression(x, y):
            """Fast linear regression using Numba.
            Parameters
            ----------
            x, y : ndarray, shape (n_times,)
                Variables
            Returns
            -------
            slope : float
                Slope of 1D least-square regression.
            intercept : float
                Intercept
            """
            n_times = x.size
            sx2 = 0
            sx = 0
            sy = 0
            sxy = 0
            for j in range(n_times):
                sx2 += x[j] ** 2
                sx += x[j]
                sxy += x[j] * y[j]
                sy += y[j]
            den = n_times * sx2 - (sx ** 2)
            num = n_times * sxy - sx * sy
            slope = num / den
            intercept = np.mean(y) - slope * np.mean(x)
            return slope, intercept

        def _higuchi_fd(x, kmax):
            n_times = x.size
            lk = np.empty(kmax)
            x_reg = np.empty(kmax)
            y_reg = np.empty(kmax)
            for k in range(1, kmax + 1):
                lm = np.empty((k,))
                for m in range(k):
                    ll = 0
                    n_max = floor((n_times - m - 1) / k)
                    n_max = int(n_max)
                    for j in range(1, n_max):
                        ll += abs(x[m + j * k] - x[m + (j - 1) * k])
                    ll /= k
                    ll *= (n_times - 1) / (k * n_max)
                    lm[m] = ll
                # Mean of lm
                m_lm = 0
                for m in range(k):
                    m_lm += lm[m]
                m_lm /= k
                lk[k - 1] = m_lm
                x_reg[k - 1] = log(1. / k)
                y_reg[k - 1] = log(m_lm)
            higuchi, _ = _linear_regression(x_reg, y_reg)
            return higuchi

        from openpyxl import Workbook
        from openpyxl.chart import (ScatterChart, Reference, Series)
        import openpyxl

        # dest_filename = 'Subj08.xlsx'
        if len(output_filename_choice.get())==0:
            dest_filename = filename +'.xlsx'
        else:
            dest_filename = output_filename_choice.get()+'.xlsx'

        if overwrite_choice == 0:
            book = Workbook()
        else:
            book = openpyxl.load_workbook(dest_filename)
        sheet = book.active
        sheet.title = 'Data'

        sheet2 = book.create_sheet()
        sheet2.title = 'Graphs'
        for h in range(len(filename)):
            temp = []
            with open(filename[h], 'r') as csvfile:
                csvreader = csv.reader(csvfile, delimiter=',')
                for row2 in csvreader:
                    if csvreader.line_num == 3:
                        temp.append(row2)
                    if csvreader.line_num >= 6:
                        if row2:
                            temp.append(row2)
                        else:
                            break

            df = pd.DataFrame(temp)  # turns the array into a dataframe
            df.columns = df.iloc[0]  # sets the column names as the first row
            df = df.drop(0)  # drops the first row since it is now a duplicate of the column names
            df.reindex(df.index.drop(1))
            df.reset_index(drop=True, inplace=True)
            df.columns = ['frames', 'subframes', 'blank', 'emg1', 'emg2', 'emg3', 'emg4', 'emg5', 'emg6', 'emg7', 'emg8',
                      'emg9', 'emg10', 'emg11', 'emg12', 'emg13', 'emg14', 'emg15', 'emg16', 'unused7', 'unused8',
                      'blank2']
            df2 = df.drop(['frames', 'subframes', 'blank', 'unused7', 'unused8', 'blank2'], axis=1)
            df2 = df2.astype(np.float)
            print(len(df))
            hor = np.arange(0, (len(df) - 0.5) / samplerate, 1 / samplerate)  # getting the time domain in seconds

            emg = []
            for i in range(0,len(emg_choice)):
                j = emg_choice[i]
                print(j)
                emg.append(df2['emg{}'.format(j)])

            cutoff_freq = 20  # ~20 Hz for movement according to emg book "Electromyography: Physiology, Engineering, and Noninvasive Apps"
            cut = cutoff_freq / nyq
            b, a = signal.butter(5, cut, btype='highpass', analog=False)
            emg_high = []
            for i in range(0, len(emg)):
                temp = signal.filtfilt(b, a, emg[i])
                emg_high.append(temp)
            # print(emg_high[0])
            # print(len(emg_high))
            # exit()

            cutoff_freq = 400  # ~500 Hz according to the emg book
            cut = cutoff_freq / nyq
            b, a = signal.butter(5, cut, btype='lowpass', analog=False)
            emg_filt = []
            for i in range(0, len(emg_high)):
                temp = signal.filtfilt(b, a, emg_high[i])
                emg_filt.append(temp)


            lowcut = 10
            highcut = 400
            # row_buff_val =
            row_buff_val = 2 * (k_val-(kmin-1))
            for t in range(len(emg_filt)):
                col = 1
                row2 = 2 + t*row_buff_val
                row = 1 + t*row_buff_val
                sheet.cell(row=row, column=col).value = 'EMG ' + str(emg_choice[t])
                # row = 2 + t*row_buff_val
                row = row + 1
                # sheet.cell(row=row, column=col).value = 1
                cellcheck = 0
                samp_len = len(emg_filt[t])
                q1_2 = floor(0.1 * samp_len)
                overlap = floor(0.3 * q1_2)
                print((len(emg_filt[t])-q1_2) // (overlap))
                # exit()
                for k_iter in range(kmin,(k_val+1)):
                    fractemp = []
                    row = row + 1
                    col = 1
                    sheet.cell(row=row, column=col).value = 'Kmax: ' + str(k_iter)

                    for v in range((len(emg_filt[t])-q1_2) // overlap):
                    # q1 = emg_filt[t]

                        q2 = emg_filt[t]
                        emg_sec = q2[(0 + (overlap * v)):(q1_2 + (overlap * v))]
                        fractemp.append(_higuchi_fd(emg_sec, floor(k_iter)))
                    # frac1 = _higuchi_fd(emg_sec, k_val)
                    # frac1 = _higuchi_fd(emg_sec, 6)
                    while cellcheck == 0:
                        if sheet.cell(row=row, column=col).value != None:
                            col = col + 1
                        else:
                            cellcheck = 1
                    frac1 = np.average(fractemp)
                    sheet.cell(row=row2, column=col).value = (col-1)
                    sheet.cell(row= row, column=col).value = frac1
                    col = col + 1

                    cellcheck = 0


            print(dest_filename)
            book.save(filename=dest_filename)

            gen_txt.configure(text='Generation Complete!', background = 'green')

    else:
        change_color()


def filebrowse():
    global csv_checker
    global filename
    global filenamesplit
    gen_txt.configure(text='', background=bg_color)
    output_filename_choice.config(state='normal')
    kval.config(state='normal')

    filename = filedialog.askopenfilenames(filetypes = (("Comma Separated Values","*.csv"),("all files","*.*")))
    for w in range(len(filename)):
        if filename[w].endswith('.csv'):
            csv_checker = 0

        else:
            file_txt.configure(text='Please choose a csv file!')
            csv_checker = 1
            break

    if csv_checker == 0:
        filenamesplitfirst = filename[0].split('/')
        filenamesplitlast = filename[-1].split('/')
        filenamesplit1 = filenamesplitfirst[-1].split('_')
        filenamesplit2 = filenamesplitlast[-1].split('_')
        trial_numberfirst = filenamesplit1[-1].split('.')
        trial_numberlast = filenamesplit2[-1].split('.')
        file_txt.configure(text='Chosen Subject: ' + filenamesplit2[1] + '\nChosen TR #s: ' + trial_numberfirst[0] + ' - ' + trial_numberlast[0])
        file_txt.config(background=bg_color)

def get_emg_vals():
    global emg_choice
    # print(emg_choice)
    emg_choice=[]
    for i in range(1,17):
        # print(i)
        # print(d['v{}'.format(i)].get())
        if d['v{}'.format(i)].get() == 1:
            emg_choice.append(i)
    print(emg_choice)

def NewFile():
    global emg_choice
    global CB1
    # global file_txt
    file_txt.configure(text='Chosen File: ')
    for i in range(1,17):
        d['v{}'.format(i)].set(0)
    gen_txt.configure(text='', background=bg_color)

def NewFileKeyboard(self):
    global emg_choice
    global CB1
    # global file_txt
    file_txt.configure(text='Chosen File: ')
    for i in range(1, 17):
        d['v{}'.format(i)].set(0)
    gen_txt.configure(text='', background=bg_color)

def ExitKeyboard(self):
    window.quit()


def Tutorial():
    top = Toplevel()
    top.title("Tutorial")
    # top.geometry('500x400')
    msg1 = Label(top, text = 'Hello!'
                            '\n Welcome to EMGProc!'
                            '\nThis software was designed to work with a 16 count EMG system, as well as'
                            '\n the csv file format created by Vicon Nexus (specifically v2.5 but possibly could'
                            '\n work for other versions)'
                            '\n'
                            '\n'
                            '\n How to use EMGProc:')
    msg1.grid(row=1,column = 1,pady = (20,0),padx = 30)

    msg2 = Label(top, text ='Step 1: Choose the csv file needed to process'
                            '\nStep 2: Choose the output filename, as well as if a new output file should be created'
                            '\n(overwriting any file with that name) or add on to an existing file created by EMGProc'
                            '\nStep 3: Choose what EMGs were active and save them'
                            '\nStep 4: Click Generate Output'
                            '\nThe output file will be placed in the same location as the chosen file, unless you choose,'
                            '\nyour own filename, in which case it will be placed in the location of this program.'
                            '\n', justify = LEFT)
    msg2.grid(row=2,column = 1,padx = 30)
    msg3 = Label(top, text ='Keyboard Shortucts:')
    msg3.grid(row=3,column = 1,padx = 30)
    msg4 = Label(top, text='Ctrl+Q: New File'
                            '\nCtrl+X: Exit', justify = LEFT)
    msg4.grid(row=4,column = 1,pady = (0,20))

    button1 = Button(top, text = 'Dismiss', command = top.destroy)
    button1.grid(row=5,column = 1,pady = (0,20))

def About():
    top = Toplevel()
    top.title("About")
    # top.geometry('500x400')
    msg1 = Label(top, text = 'EMGProc was created by Adam Lewis as part of a Master\'s Thesis Project'
                             '\nCreated with PyCharm IDE and Python 3.6'
                             '\nContact email: AdamLew73@gmail.com'
                             '\n'
                             '\nPlease do not copy or use without permission')
    msg1.grid(row=1,column = 1,pady = (20,0),padx = 30)


    button1 = Button(top, text='Dismiss', command=top.destroy)
    button1.grid(row=2, column=1, pady=(0, 20))

def focus_out_entry(event):
    output_filename.config(text='\nOutput Filename: ' + output_filename_choice.get()+'.xlsx')
def output_filename_save():
    output_filename.config(text='\nOutput Filename: ' + output_filename_choice.get() + '.xlsx')

def save_choice1():
    global overwrite_choice
    overwrite_choice = 0
    print(overwrite_choice)
def save_choice2():
    global overwrite_choice
    overwrite_choice = 1
    print(overwrite_choice)
def Pulse_Prop():
    global k_val
    k_val = int(kval.get())


##### GUI CREATION #####
window.config(menu = menu)

introtext = Label(window, text = '---------------------------------------------------------------------------------'
                               '\nFile Management')
introtext.grid(row=1,column = 2, columnspan = 3)

btn1 = Button(window, text = 'Browse for File (csv only)', command = filebrowse)
# btn1.pack(fill=X,padx = 100, pady = 50)
btn1.grid(row=2,column = 3,pady = 20)

file_txt = Label(window, text = 'Chosen File: ' + filename)
# file_txt.pack(fill=X)
file_txt.grid(row=3,column = 2, columnspan = 3, ipadx = 100)
bg_color = file_txt.cget("background")
out_text = ''
output_filename_txt = Label(window, text = 'Desired Output Filename:')
output_filename_txt.grid(row=4,column = 2, columnspan =2,padx = (25,100), sticky = W)
output_filename_choice = Entry(window, textvariable = out_text, state = 'disabled', width = 50)
output_filename_choice.bind("<Return>",focus_out_entry)
output_filename_choice.bind("<FocusOut>",focus_out_entry)
output_filename_choice.grid(row=4,column = 2, columnspan = 2, sticky = W, padx = (175,0))
output_file_save = Button(window, text = 'Enter', command = output_filename_save)
output_file_save.grid(row=4,column = 4, padx = (10,0), sticky = W)
filesavechoicevar = IntVar()
file_save_choice1 = Radiobutton(window, text = 'Overwrite File', variable = filesavechoicevar, value = 0, command = save_choice1)
file_save_choice1.grid(row=3,column = 4, padx = (50,25), pady = (10,30), rowspan = 3, sticky = NE)
file_save_choice2 = Radiobutton(window, text = 'Add to File', variable = filesavechoicevar, value = 1, command = save_choice2)
file_save_choice2.grid(row=3,column = 4, padx = (0,40), pady = (35,0), rowspan = 3, sticky = NE)

output_filename = Label(window, text = '\nOutput Filename: ')
output_filename.grid(row=5,column = 3)
septext = Label(window, text = '---------------------------------------------------------------------------------'
                               '\nActive EMGs'
                               '\n')
septext.grid(row=6,column = 2, columnspan = 3)

CB1 =Checkbutton(window, text = 'EMG 1', variable = d['v1'])
CB1.grid(row=7, column = 2, ipadx = 50)
CB2 =Checkbutton(window, text = 'EMG 2', variable = d['v2'])
CB2.grid(row=8, column = 2)
CB3 =Checkbutton(window, text = 'EMG 3', variable = d['v3'])
CB3.grid(row=9, column = 2)
CB4 =Checkbutton(window, text = 'EMG 4', variable = d['v4'])
CB4.grid(row=10, column = 2)
CB5 =Checkbutton(window, text = 'EMG 5', variable = d['v5'])
CB5.grid(row=11, column = 2)
CB6 =Checkbutton(window, text = 'EMG 6', variable = d['v6'])
CB6.grid(row=12, column = 2)
CB7 =Checkbutton(window, text = 'EMG 7', variable = d['v7'])
CB7.grid(row=7, column = 3)
CB8 =Checkbutton(window, text = 'EMG 8', variable = d['v8'])
CB8.grid(row=8, column = 3)
CB9 =Checkbutton(window, text = 'EMG 9', variable = d['v9'])
CB9.grid(row=9, column = 3)
CB10 =Checkbutton(window, text = 'EMG 10', variable = d['v10'])
CB10.grid(row=10, column = 3)
CB11 =Checkbutton(window, text = 'EMG 11', variable = d['v11'])
CB11.grid(row=11, column = 3)
CB12 =Checkbutton(window, text = 'EMG 12', variable = d['v12'])
CB12.grid(row=7, column = 4, ipadx = 50)
CB13 =Checkbutton(window, text = 'EMG 13', variable = d['v13'])
CB13.grid(row=8, column = 4)
CB14 =Checkbutton(window, text = 'EMG 14', variable = d['v14'])
CB14.grid(row=9, column = 4)
CB15 =Checkbutton(window, text = 'EMG 15', variable = d['v15'])
CB15.grid(row=10, column = 4)
CB16 =Checkbutton(window, text = 'EMG 16', variable = d['v16'])
CB16.grid(row=11, column = 4)

btn2 = Button(window, text = 'Save Active EMGs', command = get_emg_vals)
btn2.grid(row=13,column = 3,pady = 10)
# file_txt = Label(window, text = 'Chosen File: ' + filename)
# file_txt.pack(fill=X)

septext2 = Label(window, text = '---------------------------------------------------------------------------------'
                               '\nKMax Definitions'
                               '\n')
septext2.grid(row=14,column = 2, columnspan = 3)


v2 = StringVar(window, value='6')


kval_Text = Label(window, text = 'K_Max Value')
kval_Text.grid(row=15,column = 2, columnspan = 3, padx = (50,50))
kval = Entry(window, state = 'disabled', textvariable = v2, width = 10)
kval.grid(row=16,column = 2, columnspan = 3, padx = (45,50))

Pulse_Button = Button(window, text = 'Save Properties', command = Pulse_Prop)
Pulse_Button.grid(row=17,column = 3,pady = 20)

septext3 = Label(window, text = '---------------------------------------------------------------------------------')
septext3.grid(row=18,column = 2, columnspan = 3)

btn3 = Button(window, text = 'Generate Output', command = dataproc)
btn3.grid(row=19,column = 3)
gen_txt = Label(window, text = '')
gen_txt.grid(row=20,column = 3, ipadx = 150, pady = (10,20))
# gen_txt.grid(row=20,column = 2,columnspan = 3, ipadx = 150, pady = (10,20))

# btn2.grid(fill=X,padx = 100, pady = 220)


filemenu = Menu(menu, tearoff = False)
filemenu2 = Menu(menu, tearoff = False)
filemenu3 = Menu(menu, tearoff = False)

menu.add_cascade(label='File', menu = filemenu)
filemenu.add_command(label='New', command=NewFile, accelerator ="Ctrl+Q")
window.bind_all("<Control-q>",NewFileKeyboard)
filemenu.add_separator()
filemenu.add_command(label='Exit', command=window.quit)
window.bind_all("<Control-x>",ExitKeyboard)
menu.add_cascade(label='Help', menu = filemenu2)
filemenu2.add_command(label='Tutorial', command=Tutorial)

menu.add_cascade(label='Info', menu = filemenu3)
filemenu3.add_command(label='About', command=About)


window.mainloop()